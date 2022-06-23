#!/usr/bin/env python3
from os import popen, system, getcwd
from csv import reader

import openpyxl
from openpyxl import styles

XLS = []


def pre_setup() -> int:
    if not getcwd().startswith(f"/home/"):
        print(f"You might not be in a safe location. Detected path is:\n"
              f"{getcwd()}\nContinue? [y/N]", end='')
        if not input(" ") == "y":
            return -1
    print("Backing up .xlsx files to ./.temp")
    if system("mkdir .temp") != 0:
        return -1
    system("cp *.xlsx ./.temp/")
    return 0


def post_setup(undone: list) -> None:
    if undone:
        print("Some servers had multiple statuses, check them manually:")
        for i, j in enumerate(undone, 1):
            print(j, end='\n') if i % 5 == 0 else print(j, end="\t")

    if input("Finishing up, remove backups from .temp? [Y/n] ") not in {"", "\n"}:
        return
    system("rm -rf ./.temp")
    print("Finished. Remember to double check everything")


def file_setup() -> dict:
    files = {"all": list(popen("find ./ -maxdepth 1 -type f"))}
    for i in range(len(files["all"])):
        files["all"][i] = files["all"][i][2:-1]

    files["csv"] = [i for i in files["all"] if i.endswith(".csv")]
    files["xls"] = [i for i in files["all"] if i.endswith(".xlsx")]
    global XLS
    XLS = files["xls"]

    print("Report file(s) found:")
    for i in files["csv"]:
        print(i, end="\t")
    print("\n\nExcel file(s) found:")
    for i in files["xls"]:
        print(i, end="\t")
    print()
    return files


def csv_handling(files: dict) -> dict:
    servers = {}
    print("\nParsing reports...")
    for file in files["csv"]:
        with open(file, "r") as f:
            for line in reader(f, delimiter=","):
                if line[0].lower().startswith("hostname"):
                    continue

                servers[line[0]] = {}
                if line[1] != '':
                    try:
                        servers[line[0]]["status"].append(line[1])
                    except KeyError:
                        servers[line[0]]["status"] = [line[1]]
                else:
                    servers[line[0]]["status"] = []

                if line[3].lower() not in {"none required.", ""}:
                    try:
                        servers[line[0]]["comment"].append(line[3])
                    except KeyError:
                        servers[line[0]]["comment"] = [line[3]]

    # LOGIC
    for server in servers.values():
        if not server["status"] and "SERVERNOTFOUND" in server["comment"]:
            server["status"] = ["SERVERNOTFOUND"]
        if len(server["status"]) == 1:
            server["status"] = str(server["status"][0])

    print("\tdone parsing reports")
    return servers


def xls_handling(servers: dict) -> list:
    undone = []
    for file in XLS:
        spreadsheet = openpyxl.load_workbook(file)
        sheet = spreadsheet["Sheet1"]
        red = openpyxl.styles.colors.Color(rgb="00FF0000")
        green = openpyxl.styles.colors.Color(rgb="0000FF00")
        red_fill = openpyxl.styles.fills.PatternFill(patternType="solid", fgColor=red)
        green_fill = openpyxl.styles.fills.PatternFill(patternType="solid", fgColor=green)

        # letters: A - hostname, E - status, F - comments
        for row in range(1, sheet.max_row + 1):
            hostname_cell = f"A{row}"
            status_cell = f"E{row}"
            comment_cell = f"F{row}"

            hostname = sheet[hostname_cell].value

            if hostname in servers.keys():
                # print(f"{sheet[hostname].value} found in list")  # DEBUG
                if "PATCHED" in servers[hostname]["status"]:
                    sheet[status_cell] = "Successful"
                    sheet[status_cell].fill = green_fill
                else:
                    if type(servers[hostname]["status"]) is list and len(servers[hostname]["status"]) > 1:
                        undone.append(hostname)
                        continue

                    sheet[comment_cell].value = "" if sheet[comment_cell].value is None else sheet[comment_cell].value
                    sheet[status_cell] = "Unsuccessful"
                    sheet[status_cell].fill = red_fill

                    if servers[hostname]["status"] == "SKIPPED":
                        sheet[comment_cell] = "Out of scope" + sheet[comment_cell].value
                    elif servers[hostname]["status"] == "REMOVED":
                        sheet[comment_cell] = "Removed from patching" + sheet[comment_cell].value
                    elif servers[hostname]["status"] == "NOPATCHNEEDED":
                        sheet[comment_cell] = "No updates available" + sheet[comment_cell].value
                    elif servers[hostname]["status"] == "SERVERNOTFOUND":
                        sheet[comment_cell] = "Server hostname not found" + sheet[comment_cell].value

        spreadsheet.save(file)
        return undone


# CONTROL BLOCK
if __name__ == "__main__":
    if pre_setup() != 0:
        exit(-1)

    post_setup(xls_handling(csv_handling(file_setup())))
